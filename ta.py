import openpyxl
from openpyxl import load_workbook
import duckdb
import spatial
import pyfiglet 
import time     
import os
import platform


fileName='all (6).xlsx'
lang='영어' # 한글 or 영어



from datetime import datetime
now = datetime.now()
formatted = now.strftime("%Y%m%d_%H%M%S")

system_info = platform.system()  # 'Linux', 'Windows', 'Darwin'
print(system_info)

if(system_info == 'Windows'):
    os.system("cls")
else:
    os.system("clear")

result = pyfiglet.figlet_format("Trusted Advisor") 
print(result)

time.sleep(3)

if not os.path.exists('output'):import openpyxl
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")
import duckdb
import spatial
import pyfiglet
import time
import os
import platform
import pygwalker as pyg
import pandas as pd
#%matplotlib inline

fileName='all (5).xlsx'
lang='영어' # 한글 or 영어

from datetime import datetime
now = datetime.now()
formatted = now.strftime("%Y%m%d_%H%M%S")

system_info = platform.system()  # 'Linux', 'Windows', 'Darwin'
print(system_info)

if(system_info == 'Windows'):
    os.system("cls")
else:
    os.system("clear")

result = pyfiglet.figlet_format("Trusted Advisor") 
print(result)

time.sleep(3)

if not os.path.exists('output'):
    os.mkdir('output')

path = os.path.join('output','TA_')


currPath='./'


conn = duckdb.connect(database=':memory:',read_only=False)
conn.sql("""INSTALL spatial;
            LOAD spatial;
         """)

import requests

url = "https://raw.githubusercontent.com/parklikepark/Trusted_Advisor/main/TrustedAdvisorSummary.xlsx"
response = requests.get(url)

if response.status_code == 200:
    with open("TrustedAdvisorSummary_download.xlsx", "wb") as f:
        f.write(response.content)

conn.execute("""
              create table check_list as 
              select * from st_read('TrustedAdvisorSummary_download.xlsx',layer=?,open_options = ['HEADERS=FORCE', 'FIELD_TYPES=AUTO']);
              """,[lang])

conn.execute("""
              create table recommend(recomm varchar,
                                     account_id varchar,
                                     status varchar,
                                     Total_number_of_resources_processed integer,
                                     Number_of_resources_flagged integer,
                                     Number_of_suppressed_resources integer,
                                     file varchar,
                                     sheeet varchar,
                                     link varchar
                                     );
             """)			

#conn.sql("""
#            select Recommendations,count(*) as count 
#            from check_list 
#            group by Recommendations
#            order by 1
#         """).show()

wb=openpyxl.load_workbook(filename=currPath+fileName)
res=len(wb.sheetnames)
print('Number of sheets: ',res,'sheets')

ws_names=wb.sheetnames

for sn in ws_names:
    ws = wb[sn] ## 각 괄호를 이용하여 접근 가능하다.

    print('===-------------------------------------------------------===')
    print('Sheet 이름 [' + sn + ']')

    status=ws.cell(row=4,column=1).value
    res=status.split(sep=' ')[1]
    recomm=ws.cell(row=1,column=1).value
    account_id=ws.cell(row=2,column=1).value.split(sep=': ')[1]
    tot_num=ws.cell(row=6,column=2).value.split(sep=': ')[1]
    num_flagged=ws.cell(row=7,column=2).value.split(sep=': ')[1]
    num_suppressed=ws.cell(row=8,column=2).value.split(sep=': ')[1]
    link='=HYPERLINK("..\\'+fileName+'#'+'\''+sn+'\''+'!A1","Link-Click")'+'\n'
    
    conn.execute("INSERT INTO recommend VALUES (?,?,?,?,?,?,?,?,?)", [recomm, account_id,res,tot_num,num_flagged,num_suppressed,fileName,sn,link])

    
    if status  in ('Status: not_available','Status: ok','상태: not_available','상태: ok'):
        #print(ws.cell(row=1,column=1).value)
        print(status)
        continue

    for x in range(1,ws.max_row+1):
        for y in range(1,ws.max_column+1):
            if(ws.cell(row=x,column=y).value is not None):
                if(x==1 and y==1):
                   status=ws.cell(row=x,column=y).value
                   results=conn.execute("""
                                        select ifnull(min(Recommendations),'Other') as count
                                        from check_list
                                        where detail = ?
                                        """,(status,)).fetchone()

                   print('권장사항 영역 ['+results[0]+']')
                   with open(path+results[0]+'_'+formatted+'_'+res+'.txt', 'a',encoding='utf-8') as file:
                       file.write('\n=================================================================================================================\n')

                with open(path+results[0]+'_'+formatted+'_'+res+'.txt', 'a', encoding='utf-8') as file:
                    file.write(ws.cell(row=x,column=y).value+' ')

                print(ws.cell(row=x,column=y).value, end=" ")
        print()
        with open(path+results[0]+'_'+formatted+'_'+res+'.txt', 'a',encoding='utf-8') as file:
            file.write('\n')

df=conn.sql("""
            select (select max(Recommendations) from check_list cl where cl.detail=rec.recomm) recommendations,status,count(*) as count,
            sum(Total_number_of_resources_processed),
            sum(Number_of_resources_flagged),
            sum(Number_of_suppressed_resources)
            from   recommend rec
            group by Recommendations,status
            order by 5 desc,1,2
         """).to_df()
df

conn.sql("""
            select rec.recomm,Recommendations,status
            from   recommend rec left join check_list cl
            on     rec.recomm = cl.detail
            where Recommendations is null
            order by 1
         """).show()

df2=conn.sql("""
            select row_number() over( order by Number_of_resources_flagged desc,Total_number_of_resources_processed desc,recomm) SEQ, (select max(Recommendations) from check_list cl where cl.detail=rec.recomm) recommendations,
            recomm,	account_id,	status,	Total_number_of_resources_processed,	Number_of_resources_flagged, Number_of_suppressed_resources, link
            from   recommend rec
            order by Number_of_resources_flagged desc,Total_number_of_resources_processed desc,recomm
         """).to_df()
df2

writer = pd.ExcelWriter('./output/all_out_'+formatted+'.xlsx') 
df2.to_excel(writer, sheet_name='recommdations',  na_rep='NaN',float_format = "%.2f",
             header = True,
             #columns = ["group", "value_1", "value_2"], # if header is False
             index = False,
             #index_label = "id",
             #startrow = 1,
             #startcol = 1,
             #engine = 'xlsxwriter',
             freeze_panes = (1, 0))

for column in df2:
    column_width = max(df2[column].astype(str).map(len).max(), len(column))
    col_idx = df2.columns.get_loc(column)
    writer.sheets['recommdations'].set_column(col_idx, col_idx, column_width)

writer.close()


import seaborn as sns
import matplotlib.pyplot as plt

import matplotlib.font_manager as fm
font_name='Nanum'
[(f.name,f.name) for f in fm.fontManager.ttflist if f'{font_name}' in f.name]
font_path = 'NanumGothic.ttf'
fontprop = fm.FontProperties(fname=font_path)

#plt.set_title('한국어를 지정한 타이틀',fontproperties=fontprop)


plt.figure(figsize=(20, 8))
plt.title('Trusted Advisor', fontsize=20)

#sns.set_palette('twilight') 
fig=sns.barplot(data=df, x="recommendations", y="count",hue='status', ci=None)
for i in fig.containers:
    fig.bar_label(i,)

plt.legend(title='Status', fontsize='20', title_fontsize='20',prop=fontprop)

plt.xticks(rotation=0,fontsize=30,fontproperties=fontprop)

plt.show()

figure=fig.get_figure()
figure.savefig('./output/'+fileName+'_'+formatted+'_sta1.png')


plt.figure(figsize=(20, 8))
plt.title('Trusted Advisor', fontsize=20)

fig=sns.barplot(data=df, x="status", y="count",hue='recommendations', ci=None)

plt.legend(title='Recommendations', fontsize='20', title_fontsize='20',prop=fontprop)

#plt.xticks(rotation=45)
for i in fig.containers:
    fig.bar_label(i,)
    
plt.show()

figure=fig.get_figure()
figure.savefig('./output/'+fileName+'_'+formatted+'_sta2.png')


df3=df2.where(df2['Number_of_resources_flagged'] > 0)

plt.figure(figsize=(20, 50)) 
plt.title('Trusted Advisor', fontsize=20) 
plt.xticks(rotation=90)

sns.set_style('whitegrid') 
#sns.set(rc = {'figure.figsize':(20,8)})

plt.yticks(rotation=0, fontsize=30,fontproperties=fontprop)
plt.xticks(rotation=0, fontsize=20,fontproperties=fontprop)
#sns.scatterplot(data=df3, x='recomm', y='Number_of_resources_flagged')


fig=sns.barplot(data=df3, x='Number_of_resources_flagged', y='recomm',hue='status')

plt.legend(title='Status', fontsize='20', title_fontsize='20',prop=fontprop)

for i in fig.containers:
    fig.bar_label(i,)

plt.show()

figure=fig.get_figure()
figure.savefig('./output/'+fileName+'_'+formatted+'_error.png')